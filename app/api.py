import json, csv, io, os, tempfile, random
from flask import Blueprint, request, jsonify, send_file, make_response
from app.models import db, EventData
from datetime import datetime, timedelta

# For PDF generation with Platypus
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# For Excel export/import using openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

api_bp = Blueprint('api', __name__, url_prefix='/api')

def random_date(start, end):
    """Return a random datetime between start and end."""
    delta = end - start
    int_delta = delta.days
    random_days = random.randrange(int_delta + 1)
    return start + timedelta(days=random_days)

@api_bp.route('/save-event', methods=['POST'])
def save_event():
    data = request.get_json()
    try:
        # Operating Hours: default if not provided.
        if not data.get('operatingHours'):
            data['operatingHours'] = "12:00 PM - 11:00 PM"

        # Event Date Range:
        # If both eventDateFrom and eventDateTo are provided, parse them.
        # Otherwise, default both to today's date.
        if data.get('eventDateFrom') and data.get('eventDateTo'):
            event_date_from = datetime.strptime(data['eventDateFrom'], '%Y-%m-%d')
            event_date_to = datetime.strptime(data['eventDateTo'], '%Y-%m-%d')
        else:
            event_date_from = event_date_to = datetime.now()

        # Sale Hour: if not provided, use the current hour.
        if not data.get('saleHour'):
            sale_hour = datetime.now().hour
        else:
            sale_hour = int(data['saleHour'])

        # Sales Volume: if not provided, generate a random realistic value.
        if not data.get('salesVolume'):
            sales_volume = round(random.uniform(50, 500), 2)
        else:
            sales_volume = float(data['salesVolume'])

        # Price Per Unit: if not provided, generate a random value.
        if not data.get('pricePerUnit'):
            price_per_unit = round(random.uniform(1, 10), 2)
        else:
            price_per_unit = float(data['pricePerUnit'])

        # Total Revenue: computed if not provided.
        if not data.get('totalRevenue'):
            total_revenue = round(sales_volume * price_per_unit, 2)
        else:
            total_revenue = float(data['totalRevenue'])

        # Payment Method: if not provided, choose randomly.
        if not data.get('paymentMethod'):
            payment_method = random.choice(["Cash", "Card", "Contactless"])
        else:
            payment_method = data['paymentMethod']

        # Products Sold: expect a list.
        products_sold_json = json.dumps(data.get('selectedProducts', []))

        new_event = EventData(
            event_name=data['eventName'],
            venue_name=data['venueName'],
            operating_hours=data['operatingHours'],
            event_date_from=event_date_from,
            event_date_to=event_date_to,
            products_sold=products_sold_json,
            sales_volume=sales_volume,
            price_per_unit=price_per_unit,
            total_revenue=total_revenue,
            sale_hour=sale_hour,
            payment_method=payment_method
        )
        db.session.add(new_event)
        db.session.commit()
        return jsonify({"message": "Event saved successfully!"}), 201

    except Exception as e:
        return jsonify({"message": f"Error saving event: {str(e)}"}), 400

@api_bp.route('/get-events', methods=['GET'])
def get_events():
    try:
        events = EventData.query.all()
        events_data = []
        for e in events:
            events_data.append({
                "id": e.id,
                "event_name": e.event_name,
                "event_date_from": e.event_date_from.strftime('%Y-%m-%d') if e.event_date_from else "",
                "event_date_to": e.event_date_to.strftime('%Y-%m-%d') if e.event_date_to else "",
                "venue_name": e.venue_name,
                "operating_hours": e.operating_hours,
                "products_sold": e.products_sold,  # JSON string
                "sales_volume": e.sales_volume,
                "price_per_unit": e.price_per_unit,
                "total_revenue": e.total_revenue,
                "sale_hour": e.sale_hour,
                "payment_method": e.payment_method
            })
        return jsonify(events_data), 200
    except Exception as e:
        return jsonify({"message": f"Error fetching events: {str(e)}"}), 400

@api_bp.route('/import-events', methods=['POST'])
def import_events():
    if 'file' not in request.files:
        return jsonify({"message": "No file part in the request"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"message": "No file selected"}), 400

    try:
        filename = file.filename.lower()
        events_imported = 0

        if filename.endswith(".csv"):
            stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
            csv_input = csv.DictReader(stream)
            for row in csv_input:
                try:
                    products = row['selectedProducts']
                    try:
                        products_list = json.loads(products)
                    except Exception:
                        products_list = [p.strip() for p in products.split(",")]
                    new_event = EventData(
                        event_name=row['eventName'],
                        event_date_from=datetime.strptime(row['eventDateFrom'], '%Y-%m-%d'),
                        event_date_to=datetime.strptime(row['eventDateTo'], '%Y-%m-%d'),
                        venue_name=row['venueName'],
                        operating_hours=row['operatingHours'],
                        products_sold=json.dumps(products_list),
                        sales_volume=float(row['salesVolume']),
                        price_per_unit=float(row['pricePerUnit']),
                        total_revenue=float(row['totalRevenue']),
                        sale_hour=int(row['saleHour']),
                        payment_method=row['paymentMethod']
                    )
                    db.session.add(new_event)
                    events_imported += 1
                except Exception:
                    continue
        elif filename.endswith((".xlsx", ".xls")):
            in_mem_file = io.BytesIO(file.read())
            wb = load_workbook(filename=in_mem_file, read_only=True)
            sheet = wb.active
            headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            header_map = {h: i for i, h in enumerate(headers)}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                try:
                    eventName = row[header_map["eventName"]]
                    eventDateFrom = row[header_map["eventDateFrom"]]
                    eventDateTo = row[header_map["eventDateTo"]]
                    venueName = row[header_map["venueName"]]
                    operatingHours = row[header_map["operatingHours"]]
                    selectedProducts = row[header_map["selectedProducts"]]
                    salesVolume = row[header_map["salesVolume"]]
                    pricePerUnit = row[header_map["pricePerUnit"]]
                    totalRevenue = row[header_map["totalRevenue"]]
                    saleHour = row[header_map["saleHour"]]
                    paymentMethod = row[header_map["paymentMethod"]]

                    try:
                        products_list = json.loads(selectedProducts)
                    except Exception:
                        products_list = [p.strip() for p in str(selectedProducts).split(",")]

                    new_event = EventData(
                        event_name=str(eventName),
                        event_date_from=datetime.strptime(str(eventDateFrom), "%Y-%m-%d"),
                        event_date_to=datetime.strptime(str(eventDateTo), "%Y-%m-%d"),
                        venue_name=str(venueName),
                        operating_hours=str(operatingHours),
                        products_sold=json.dumps(products_list),
                        sales_volume=float(salesVolume),
                        price_per_unit=float(pricePerUnit),
                        total_revenue=float(totalRevenue),
                        sale_hour=int(saleHour),
                        payment_method=str(paymentMethod)
                    )
                    db.session.add(new_event)
                    events_imported += 1
                except Exception:
                    continue
        else:
            return jsonify({"message": "Unsupported file type"}), 400

        db.session.commit()
        return jsonify({"message": f"Successfully imported {events_imported} events."}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"message": f"Error importing events: {str(e)}"}), 400

@api_bp.route('/export-csv', methods=['GET'])
def export_csv():
    try:
        events = EventData.query.all()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "eventName", "eventDateFrom", "eventDateTo", "venueName", "operatingHours",
            "selectedProducts", "salesVolume", "pricePerUnit", "totalRevenue", "saleHour", "paymentMethod"
        ])
        for evt in events:
            writer.writerow([
                evt.event_name,
                evt.event_date_from.strftime('%Y-%m-%d') if evt.event_date_from else "",
                evt.event_date_to.strftime('%Y-%m-%d') if evt.event_date_to else "",
                evt.venue_name,
                evt.operating_hours,
                evt.products_sold,
                evt.sales_volume,
                evt.price_per_unit,
                evt.total_revenue,
                evt.sale_hour,
                evt.payment_method
            ])
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=sales_report.csv"
        response.headers["Content-type"] = "text/csv"
        return response
    except Exception as e:
        return jsonify({"message": f"Error exporting CSV: {str(e)}"}), 400

@api_bp.route('/export-excel', methods=['GET'])
def export_excel():
    try:
        events = EventData.query.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        headers = [
            "eventName", "eventDateFrom", "eventDateTo", "venueName", "operatingHours",
            "selectedProducts", "salesVolume", "pricePerUnit", "totalRevenue", "saleHour", "paymentMethod"
        ]
        ws.append(headers)
        for evt in events:
            ws.append([
                evt.event_name,
                evt.event_date_from.strftime('%Y-%m-%d') if evt.event_date_from else "",
                evt.event_date_to.strftime('%Y-%m-%d') if evt.event_date_to else "",
                evt.venue_name,
                evt.operating_hours,
                evt.products_sold,
                evt.sales_volume,
                evt.price_per_unit,
                evt.total_revenue,
                evt.sale_hour,
                evt.payment_method
            ])
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(
            tmp.name,
            as_attachment=True,
            download_name="sales_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"message": f"Error exporting Excel: {str(e)}"}), 400
@api_bp.route('/export-pdf', methods=['GET'])
def export_pdf():
    """
    Generates a PDF with:
      - A two-column header row: logo (left) and text (title + filters) (right)
      - A data table for events, with random placeholder values for missing fields
      - A "Grand Total" row at the bottom of the table
      - A mini-table containing pie chart and bar chart, centered
      - A footer timestamp
    """
    try:
        import os
        import json
        import random
        import tempfile
        from datetime import datetime
        from collections import defaultdict

        import matplotlib.pyplot as plt
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        # Hard-coded "header" info
        date_range_str = "01/05/2023 - 01/05/2024"
        report_filters_str = "Dept: Sales"
        site_str = "0 Head Office"

        # Fetch events
        events = EventData.query.all()
        if not events:
            return jsonify({"message": "No events found to generate report."}), 400

        # Functions for random placeholders if data is missing
        def random_price():
            return round(random.uniform(1, 10), 2)

        def random_volume():
            return round(random.uniform(1, 500), 2)

        def random_product():
            return random.choice(["Fosters", "Amstel", "Heineken", "Cruzcampo", "Budweiser", "Guinness"])

        def random_date_str():
            # random date in 2025
            day = random.randint(1, 28)
            month = random.randint(1, 12)
            return f"2025-{month:02d}-{day:02d}"

        # Build table data
        table_data = [["Event Name", "Date", "Prod Name", "Sales Vol", "Price/Unit"]]

        # For chart data
        product_volume_map = defaultdict(float)
        daily_revenue_map = defaultdict(float)
        grand_total = 0.0

        for evt in events:
            # 1) Event Name
            event_name = evt.event_name if evt.event_name else f"Event-{random.randint(100,999)}"
            # 2) Date
            if evt.event_date_from:
                date_str = evt.event_date_from.strftime("%Y-%m-%d")
            else:
                date_str = random_date_str()
            # 3) Products
            try:
                products_list = json.loads(evt.products_sold) or []
            except:
                products_list = []
            if not products_list:
                products_list = [random_product()]
            products_str = ", ".join(products_list)
            # 4) Sales Volume
            vol = evt.sales_volume if evt.sales_volume else random_volume()
            # 5) Price per Unit
            ppu = evt.price_per_unit if evt.price_per_unit else random_price()

            # Update chart data
            product_volume_map[products_list[0]] += vol
            daily_revenue_map[date_str] += (vol * ppu)

            # Grand total
            row_total = vol * ppu
            grand_total += row_total

            # Add row to table
            table_data.append([
                event_name,
                date_str,
                products_str,
                f"{vol:.2f}",
                f"${ppu:.2f}"
            ])

        # Grand total row
        table_data.append(["", "", "", "Grand Total:", f"${grand_total:.2f}"])

        # Generate Pie Chart (product distribution by volume)
        from io import BytesIO
        pie_buf = BytesIO()
        product_labels = list(product_volume_map.keys())
        product_volumes = list(product_volume_map.values())
        if sum(product_volumes) == 0:
            product_labels = ["No Data"]
            product_volumes = [1]
        plt.figure(figsize=(3, 3))
        plt.pie(product_volumes, labels=product_labels, autopct='%1.1f%%', startangle=140)
        plt.title("Product Distribution", fontsize=10)
        plt.tight_layout()
        plt.savefig(pie_buf, format='png')
        plt.close()
        pie_buf.seek(0)

        # Generate Bar Chart (daily revenue)
        bar_buf = BytesIO()
        daily_dates = sorted(daily_revenue_map.keys())
        daily_values = [daily_revenue_map[d] for d in daily_dates]
        if not daily_dates:
            daily_dates = ["No Data"]
            daily_values = [1]
        plt.figure(figsize=(4, 2.5))
        plt.bar(daily_dates, daily_values, color='#0d6efd')
        plt.title("Daily Sales (Revenue)", fontsize=10)
        plt.xlabel("Date", fontsize=8)
        plt.ylabel("Revenue ($)", fontsize=8)
        plt.xticks(rotation=45, fontsize=6)
        plt.yticks(fontsize=6)
        plt.tight_layout()
        plt.savefig(bar_buf, format='png')
        plt.close()
        bar_buf.seek(0)

        # Build PDF
        import tempfile
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        doc = SimpleDocTemplate(
            tmp_file.name,
            pagesize=A4,
            topMargin=40, bottomMargin=40,
            leftMargin=30, rightMargin=30
        )
        styles = getSampleStyleSheet()
        flowables = []

        # Build a mini-table for the top row: logo on the left, text on the right
        from reportlab.platypus import Table as RLTable, TableStyle as RLTableStyle

        # (A) Logo
        logo_path = "app/static/logo.png"
        if os.path.exists(logo_path):
            logo_img = Image(logo_path, width=40, height=40)
        else:
            # Fallback if no logo
            logo_img = Paragraph("<b>No Logo</b>", styles["Normal"])

        # (B) Header text: Title + date range + filters + site
        header_paras = []
        header_paras.append(Paragraph("<b>Event Sales Report</b>", styles["Title"]))
        header_paras.append(Spacer(1, 4))
        header_paras.append(Paragraph(f"Date Range: {date_range_str}", styles["Normal"]))
        header_paras.append(Paragraph(f"Report Filters: {report_filters_str}", styles["Normal"]))
        header_paras.append(Paragraph(f"Site: {site_str}", styles["Normal"]))

        header_data = [[logo_img, header_paras]]
        header_table = RLTable(header_data, colWidths=[50, 400])
        header_table.setStyle(RLTableStyle([
            ("VALIGN", (0,0), (0,0), "TOP"),
            ("VALIGN", (1,0), (1,0), "TOP"),
            ("LEFTPADDING", (0,0), (-1,-1), 0),
            ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ]))

        flowables.append(header_table)
        flowables.append(Spacer(1, 12))

        # (C) Data Table
        col_widths = [80, 60, 140, 60, 60]
        data_table = Table(table_data, colWidths=col_widths, hAlign="LEFT", repeatRows=1)
        data_table.setStyle(TableStyle([
            ("BOX", (0,0), (-1,-1), 1, colors.black),
            ("INNERGRID", (0,0), (-1,-2), 0.5, colors.grey),
            ("INNERGRID", (0,-1), (-1,-1), 0.5, colors.black),
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("ALIGN", (3,1), (4,-1), "RIGHT"),  # numeric columns right aligned
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("TOPPADDING", (0,0), (-1,-1), 4),
        ]))
        flowables.append(data_table)
        flowables.append(Spacer(1, 12))

        # (D) Charts side by side, centered
        pie_img = Image(pie_buf, width=150, height=150)
        bar_img = Image(bar_buf, width=200, height=120)
        charts_data = [[pie_img, bar_img]]
        charts_table = RLTable(charts_data, colWidths=[170, 220])
        charts_table.setStyle(RLTableStyle([
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ]))
        flowables.append(charts_table)
        flowables.append(Spacer(1, 12))

        # (E) Footer with timestamp
        from reportlab.platypus import HRFlowable
        flowables.append(HRFlowable(width="100%", color=colors.black, thickness=1))
        flowables.append(Spacer(1, 6))
        timestamp_str = datetime.now().strftime("Receipt Generated: %d/%m/%Y %H:%M:%S")
        flowables.append(Paragraph(timestamp_str, styles["Normal"]))

        doc.build(flowables)

        return send_file(
            tmp_file.name,
            as_attachment=True,
            download_name="event_sales_receipt.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        return jsonify({"message": f"Error exporting PDF: {str(e)}"}), 400
